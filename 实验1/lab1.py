# coding=utf-8
# author：dzzhyk
# date：2020年10月25日16:33:48
# title：lab1.py

from pulp import *
from openpyxl import load_workbook
import numpy as np


NODE_LIST = []              # 节点列表
ORIGIN_NODE_GRAPH = {}      # 原始图
CHILD_NUMBER = 4            # 最大子节点数量
PE = 16000                     # PE数量
NODE_COUNT = 0              # 节点数量
T = 0                       # 最大时间步
II = 10000                      # II数量


# 节点类
class Node:
    def __init__(self, _id, child_list, t_start, t_end, node_type, has_parent):
        self.now_id = _id               # 现节点编号
        self.origin_id = _id            # 原节点编号
        self.child_list = child_list    # 子节点列表
        self.t_start = t_start          # 开始
        self.t_end = t_end              # 结束
        self.node_type = node_type      # 节点类型
        self.has_parent = has_parent    # 是否有父节点
        self.t_step = t_start           # 时间步初始化为t_start

    def __str__(self):
        return f"[Node] now_id={self.now_id}, origin_id={self.origin_id}, child_list=[" + ",".join(
            str(c) for c in self.child_list) + \
               f"], t_start={self.t_start}, t_end={self.t_end}, node_type={self.node_type}, has_parent={self.has_parent}\n"

    def toFile(self):
        return f"{self.now_id},{self.t_step},{','.join([str(c) for c in self.child_list])}, {self.node_type}, {self.origin_id}\n"


# 读入数据
def read(filename):
    # 更新相关常数
    global NODE_COUNT, T
    with open(filename, "r", encoding="utf-8") as fin:
        print("<=== 读入文件：" + filename)
        temp = fin.readline()
        while temp != "":
            temp_ls = temp.split(",")
            temp_id = temp_ls[0]
            temp_child = [int(temp_ls[i + 1]) for i in range(CHILD_NUMBER)]
            temp_start, temp_end, temp_type, temp_parent =\
                [int(a) for a in temp_ls[CHILD_NUMBER + 1:CHILD_NUMBER + 5]]
            NODE_LIST.append(Node(temp_id, temp_child, temp_start, temp_end, temp_type, temp_parent))

            # 更新全局常数
            NODE_COUNT += 1
            T = max(T, temp_end+1)
            temp = fin.readline()


# 读取xlsx
def read_xlsx(dirname, filename):
    # 更新相关常数
    global NODE_COUNT, T
    # 加载xlsx文件
    workbook = load_workbook(dirname + filename)
    sheet = workbook.active  # 获取sheet表
    # 获取sheet页的行数据
    rows = sheet.rows
    # 获取sheet页的列数据
    columns = sheet.columns
    i = 6
    # 迭代所有的行
    for row in rows:
        i += 1
        temp_ls = []
        if sheet.cell(row=i, column=1).value is None:
            break
        for j in range(1, 10):
            temp_ls.append(str(sheet.cell(row=i, column=j).value))
        temp_id = temp_ls[0]
        temp_child = [int(temp_ls[i + 1]) for i in range(CHILD_NUMBER)]
        temp_start, temp_end, temp_type, temp_parent = \
            [int(a) for a in temp_ls[CHILD_NUMBER + 1:CHILD_NUMBER + 5]]
        NODE_LIST.append(Node(temp_id, temp_child, temp_start, temp_end, temp_type, temp_parent))

        # 更新全局常数
        NODE_COUNT += 1
        T = max(T, temp_end + 1)

        print(" ".join(temp_ls))


# 写入xlsx
def write_xlsx(dirname, filename):

    pass


# 输出结果
def write(filename):
    with open(filename, "w", encoding="utf-8") as fout:
        print("===> 输出目标：" + filename)
        for node in NODE_LIST:
            fout.write(node.toFile())


# 预处理输入数据
def pre():
    for node in NODE_LIST:
        ORIGIN_NODE_GRAPH[str(node.origin_id)] = [str(c) for c in node.child_list if c != 0]


# 打印图
def printG(G):
    for dic in G:
        print(dic, end=": ")
        print(G[dic])


# 求解
def ILP(pe, points, times, G, ii):
    prob = LpProblem("lab1", LpMinimize)

    PE = LpVariable("PE", lowBound=0, upBound=pe, cat="Integer")
    prob += PE

    vars = [LpVariable("x" + str(i) + "_" + str(k), lowBound=0, upBound=1, cat="Integer")
            for i in range(points) for k in range(times)]

    print(f"=========================定义变量=========================\n{vars}")

    for i in range(points):
        prob += (lpSum(vars[i * 8: i * 8 + times]) == 1)
    vars = np.array(vars)

    for i in G:
        values = G[i]
        for v in values:
            node_child = []
            node_father = []
            for t in range(times):
                node_child.append(vars[(int(v) - 1) * 8 + t] * t)
                node_father.append(vars[(int(i) - 1) * 8 + t] * t)
            prob += (lpSum(node_child) >= lpSum(node_father) + 1)

    for j in range(times):
        pe_cons = []
        for i in range(points):
            pe_cons.append(vars[i * 8 + j])
        prob += (lpSum(pe_cons) <= PE)

    # 循环次数
    maxListNum = np.math.ceil(times / ii)
    for i in range(ii):
        temp_ii_cons = []
        for k in range(maxListNum):
            if (i + k * ii) < times:
                for p in range(points):
                    temp_ii_cons.append(vars[p * times + i + k * ii])
        prob += (lpSum(temp_ii_cons) <= PE)

    # 打印所有约束
    print(f"=========================约束信息=========================\n{prob}")

    # 求解
    print("=========================进行求解=========================")
    prob.solve()

    print("=========================输出结果=========================")
    for i in range(points):
        timeList = []
        for j in range(times):
            timeList.append(j * vars[i * 8 + j].varValue)
        temp_step = int(sum(timeList))
        # 把结果添加到NODE_LIST中的节点信息中
        NODE_LIST[i].t_step = temp_step
        print(f"{i + 1} 节点时间步是 {temp_step}")


def solve():

    # 读入数据
    # read("./in")
    read_xlsx("./dfg/", "g3.xlsx")

    # 预处理得到精简图
    pre()

    # 打印精简图信息
    print("=========================获取图信息=========================")
    printG(ORIGIN_NODE_GRAPH)

    # 求解
    ILP(PE, NODE_COUNT, T, ORIGIN_NODE_GRAPH, II)

    # 输出
    write("./out")


if __name__ == "__main__":
    solve()
    # read_xlsx("./dfg/", "g3.xlsx")
    # pre()
    # print("=========================获取图信息=========================")
    # printG(ORIGIN_NODE_GRAPH)